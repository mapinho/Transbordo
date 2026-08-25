from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.simulacao.planilha import (
    ABAS_NA_ORDEM, COLUNAS_POR_ABA, analisar, gerar_template,
)


class GerarTemplateTests(SimpleTestCase):
    def test_tem_as_cinco_abas_na_ordem_de_dependencia(self):
        wb = load_workbook(gerar_template())

        self.assertEqual(wb.sheetnames, ABAS_NA_ORDEM)

    def test_cada_aba_tem_o_cabecalho_que_o_parser_espera(self):
        wb = load_workbook(gerar_template())

        for nome in ABAS_NA_ORDEM:
            cabecalho = [c.value for c in wb[nome][1]]
            self.assertEqual(cabecalho, COLUNAS_POR_ABA[nome], f'aba {nome}')

    def test_nao_traz_linhas_de_dados(self):
        wb = load_workbook(gerar_template())

        for nome in ABAS_NA_ORDEM:
            self.assertEqual(wb[nome].max_row, 1, f'aba {nome}')

    def test_o_proprio_parser_aceita_o_template_vazio(self):
        """A prova de que template e parser não divergem."""
        relatorio = analisar(gerar_template(), None)

        self.assertFalse(relatorio.tem_erro_estrutural)
        self.assertEqual(relatorio.total_criar, 0)
