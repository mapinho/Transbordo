import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.core.models import Cooperativa
from apps.simulacao.models import Armazem, Cenario, ResumoMensalArmazem

User = get_user_model()


class ExportTests(TestCase):
    def setUp(self):
        self.coop = Cooperativa.objects.create(nome="C", slug="c")
        self.user = User.objects.create_user(
            username="u", email="u@t.test", password="x",
            papel=User.PAPEL_ADMIN_COOPERATIVA, cooperativa=self.coop)
        self.cen = Cenario.all_cooperativas.create(cooperativa=self.coop, nome="Cen")
        arm = Armazem.all_cooperativas.create(
            cooperativa=self.coop, cenario=self.cen, nome="A",
            capacidade_estatica=1, capacidade_expedicao_diaria=1, estoque_inicial=0)
        for mes in ("2026-01", "2026-02", "2026-03"):
            ResumoMensalArmazem.all_cooperativas.create(
                cooperativa=self.coop, cenario=self.cen, armazem=arm, mes=mes,
                rec_produtor=10, envio_transbordo=2, vendas=1, saldo_estoque=5,
                capacidade_estatica=200, excedente=0)
        self.url = reverse("simulacao:estoque_export", kwargs={"cenario_id": self.cen.id})

    def test_xlsx_conteudo_e_headers(self):
        self.client.force_login(self.user)
        r = self.client.get(self.url, {"visao": "armazem", "formato": "xlsx"})
        self.assertEqual(r.status_code, 200)
        self.assertIn("attachment", r["Content-Disposition"])
        conteudo = b"".join(r.streaming_content) if hasattr(r, "streaming_content") else r.content
        ws = load_workbook(io.BytesIO(conteudo)).active
        self.assertEqual(ws.cell(1, 1).value, "Mês")
        self.assertEqual(ws.max_row, 4)   # header + 3 meses
        self.assertIsInstance(ws.cell(2, 3).value, (int, float))   # rec_produtor numérico

    def test_csv_ponto_e_virgula_e_bom(self):
        self.client.force_login(self.user)
        r = self.client.get(self.url, {"visao": "sistema", "formato": "csv"})
        conteudo = b"".join(r.streaming_content) if hasattr(r, "streaming_content") else r.content
        self.assertTrue(conteudo.startswith(b"\xef\xbb\xbf"))
        self.assertIn(";", conteudo.decode("utf-8-sig").splitlines()[0])

    def test_formato_invalido_400(self):
        self.client.force_login(self.user)
        self.assertEqual(self.client.get(self.url, {"formato": "pdf"}).status_code, 400)

    def test_export_recorte_grande_400(self):
        from apps.simulacao import estoque
        orig = estoque.EXPORT_MAX
        estoque.EXPORT_MAX = 1
        try:
            self.client.force_login(self.user)
            r = self.client.get(self.url, {"visao": "armazem", "formato": "csv"})
            self.assertEqual(r.status_code, 400)
        finally:
            estoque.EXPORT_MAX = orig

    def test_gate_admin_vector_sem_org(self):
        v = User.objects.create_user(username="v", email="v@t.test", password="x",
                                     papel=User.PAPEL_ADMIN_VECTOR)
        self.client.force_login(v)
        self.assertEqual(self.client.get(self.url, {"formato": "csv"}).status_code, 403)

    def test_anonimo_redireciona_login(self):
        self.assertIn("/accounts/login/", self.client.get(self.url, {"formato": "csv"}).url)
