"""Monta pastas .xlsx em memória para os testes de planilha.

Existe para que os testes do parser não precisem de arquivo em disco nem de
fixture versionada -- cada teste declara exatamente as linhas de que precisa.
"""
import io

from openpyxl import Workbook

from apps.simulacao.planilha import (
    ABA_ARMAZENS, ABA_FABRICAS, ABA_PREVISOES, ABA_ROTAS, ABA_SAFRAS,
)

NOME_DA_ABA = {
    'fabricas': ABA_FABRICAS,
    'armazens': ABA_ARMAZENS,
    'rotas': ABA_ROTAS,
    'previsoes': ABA_PREVISOES,
    'safras': ABA_SAFRAS,
}


def montar_pasta(**abas):
    """montar_pasta(fabricas=[{'nome': 'X', ...}, ...]) -> BytesIO.

    A ordem das chaves do primeiro dict de cada aba define o cabeçalho.
    """
    wb = Workbook()
    wb.remove(wb.active)
    for chave, linhas in abas.items():
        ws = wb.create_sheet(NOME_DA_ABA[chave])
        if not linhas:
            continue
        colunas = list(linhas[0].keys())
        ws.append(colunas)
        for linha in linhas:
            ws.append([linha.get(c) for c in colunas])
    return _salvar(wb)


def montar_pasta_bruta(abas):
    """abas = {'Fábricas': [['nome', 'x'], ['A', 1]]} -- listas de listas, para
    exercitar cabeçalho irreconhecível e outros casos estruturais."""
    wb = Workbook()
    wb.remove(wb.active)
    for nome, linhas in abas.items():
        ws = wb.create_sheet(nome)
        for linha in linhas:
            ws.append(linha)
    return _salvar(wb)


def _salvar(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
