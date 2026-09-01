import csv
import datetime
import io
from unittest import mock

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.core.models import Cooperativa
from apps.simulacao import resultados
from apps.simulacao.models import Armazem, Cenario, Fabrica, MovimentacaoDiaria

User = get_user_model()


class ExportTests(TestCase):
    def setUp(self):
        self.coop = Cooperativa.objects.create(nome="C", slug="c")
        self.user = User.objects.create_user(
            username="u", email="u@t.test", password="x",
            papel=User.PAPEL_ADMIN_COOPERATIVA, cooperativa=self.coop)
        self.cen = Cenario.all_cooperativas.create(cooperativa=self.coop, nome="Cen")
        arm = Armazem.all_cooperativas.create(
            cooperativa=self.coop, cenario=self.cen, nome="A",
            capacidade_estatica=1, capacidade_expedicao_diaria=1, estoque_inicial=0)
        fab = Fabrica.all_cooperativas.create(
            cooperativa=self.coop, cenario=self.cen, nome="F",
            capacidade_estatica=1, capacidade_esmagamento_diaria=1,
            capacidade_recebimento_diaria=1, limite_caminhoes=1,
            carga_media_caminhao=1, estoque_inicial=0)
        for i in range(3):
            MovimentacaoDiaria.all_cooperativas.create(
                cooperativa=self.coop, cenario=self.cen,
                data=datetime.date(2026, 1, 5 + i), armazem=arm, fabrica=fab,
                quantidade_ton=10, custo_total=100)
        self.url = reverse("simulacao:resultados_export", kwargs={"cenario_id": self.cen.id})

    def test_xlsx_conteudo_e_headers(self):
        self.client.force_login(self.user)
        r = self.client.get(self.url, {"periodo": "diario", "agrupar": "fabrica_armazem",
                                       "formato": "xlsx"})
        self.assertEqual(r.status_code, 200)
        self.assertIn("attachment", r["Content-Disposition"])
        self.assertIn(".xlsx", r["Content-Disposition"])
        wb = load_workbook(io.BytesIO(b"".join(r.streaming_content) if hasattr(r, "streaming_content") else r.content))
        ws = wb.active
        self.assertEqual(ws.cell(1, 1).value, "Dia")
        self.assertEqual(ws.max_row, 4)  # 1 header + 3 linhas
        self.assertIsInstance(ws.cell(2, 4).value, (int, float))  # ton = número

    def test_csv_ponto_e_virgula_e_bom(self):
        self.client.force_login(self.user)
        r = self.client.get(self.url, {"formato": "csv"})
        conteudo = (b"".join(r.streaming_content) if hasattr(r, "streaming_content") else r.content)
        self.assertTrue(conteudo.startswith(b"\xef\xbb\xbf"))
        texto = conteudo.decode("utf-8-sig")
        self.assertIn(";", texto.splitlines()[0])

    def test_formato_invalido_400(self):
        self.client.force_login(self.user)
        self.assertEqual(self.client.get(self.url, {"formato": "pdf"}).status_code, 400)

    def test_respeita_filtro(self):
        self.client.force_login(self.user)
        r = self.client.get(self.url, {"formato": "csv", "data_de": "2026-01-06"})
        conteudo = (b"".join(r.streaming_content) if hasattr(r, "streaming_content") else r.content)
        self.assertEqual(len(conteudo.decode("utf-8-sig").strip().splitlines()), 3)  # header + 2

    def test_export_recorte_grande_400(self):
        self.client.force_login(self.user)
        with mock.patch.object(resultados, "EXPORT_MAX", 2):
            r = self.client.get(self.url, {"formato": "csv", "periodo": "diario",
                                           "agrupar": "fabrica_armazem"})
        self.assertEqual(r.status_code, 400)

    def test_export_comparar_nao_numerico_nao_quebra(self):
        self.client.force_login(self.user)
        r = self.client.get(self.url, {"formato": "csv", "comparar": "abc"})
        self.assertEqual(r.status_code, 200)

    def test_export_anonimo_redireciona_login(self):
        r = self.client.get(self.url, {"formato": "csv"})
        self.assertEqual(r.status_code, 302)
        self.assertIn("/accounts/login/", r["Location"])

    def test_gate_admin_vector_sem_org(self):
        v = User.objects.create_user(username="v", email="v@t.test", password="x",
                                     papel=User.PAPEL_ADMIN_VECTOR)
        self.client.force_login(v)
        self.assertEqual(self.client.get(self.url, {"formato": "csv"}).status_code, 403)
