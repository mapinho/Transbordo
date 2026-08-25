"""Importação de dados de cenário a partir de uma pasta .xlsx de cinco abas.

`analisar` lê e classifica sem escrever nada; `aplicar` grava. A fronteira
existe para que a pré-visualização seja confiável e para que quase toda a
lógica seja testável com uma pasta montada em memória. Ver
docs/superpowers/specs/2026-08-25-carga-de-dados-design.md.
"""
import datetime
from dataclasses import dataclass, field

from django.db import transaction
from openpyxl import load_workbook

from apps.simulacao.models import (
    Armazem, Cenario, Fabrica, PrevisaoArmazem, PrevisaoFabrica, Rota, SafraUnidade,
)

ABA_FABRICAS = 'Fábricas'
ABA_ARMAZENS = 'Armazéns'
ABA_ROTAS = 'Rotas'
ABA_PREVISOES = 'Previsões'
ABA_SAFRAS = 'Safras'

# A ordem É a ordem de dependência: Rotas, Previsões e Safras resolvem nomes
# contra o que Fábricas e Armazéns criaram antes delas.
ABAS_NA_ORDEM = [ABA_FABRICAS, ABA_ARMAZENS, ABA_ROTAS, ABA_PREVISOES, ABA_SAFRAS]

COLUNAS_POR_ABA = {
    ABA_FABRICAS: [
        'nome', 'capacidade_estatica', 'capacidade_esmagamento_diaria',
        'capacidade_recebimento_diaria', 'limite_caminhoes',
        'carga_media_caminhao', 'estoque_inicial',
    ],
    ABA_ARMAZENS: [
        'nome', 'capacidade_estatica', 'capacidade_expedicao_diaria', 'estoque_inicial',
    ],
    ABA_ROTAS: [
        'origem', 'destino', 'distancia_km', 'custo_frete_ton', 'custo_frete_entressafra',
    ],
    ABA_PREVISOES: ['entidade', 'mes_referencia', 'recebimento_produtor', 'vendas'],
    ABA_SAFRAS: ['unidade', 'data_inicio', 'data_fim'],
}

# Numéricos obrigatórios: célula em branco rejeita a linha. Espelha
# FABRICA/ARMAZEM_CAMPOS_NUMERICOS_OBRIGATORIOS do data_loader.py legado -- é a
# correção do bug A8 da Fase 1, em que NaN do pandas chegava ao Postgres como
# número válido.
OBRIGATORIOS_POR_ABA = {
    ABA_FABRICAS: COLUNAS_POR_ABA[ABA_FABRICAS][1:],
    ABA_ARMAZENS: COLUNAS_POR_ABA[ABA_ARMAZENS][1:],
}

# Fábricas e Armazéns são as únicas abas cuja coluna 'nome' é gravada direto
# num CharField do model (Rotas/Previsões/Safras só usam o texto da célula
# para resolver contra essas duas -- não persistem texto livre). O limite
# vem do próprio model para não arriscar divergir de models.py (finding 2 do
# review da Task 3).
MODELO_POR_ABA = {ABA_FABRICAS: Fabrica, ABA_ARMAZENS: Armazem}


@dataclass
class LinhaRejeitada:
    aba: str
    linha: int  # número como aparece no Excel: o cabeçalho é 1
    motivo: str
    valores: dict


@dataclass
class ResumoAba:
    aba: str
    criar: int = 0
    atualizar: int = 0
    rejeitadas: list = field(default_factory=list)


@dataclass
class Relatorio:
    abas: list = field(default_factory=list)
    erro_estrutural: str = None

    @property
    def tem_erro_estrutural(self):
        return self.erro_estrutural is not None

    @property
    def total_criar(self):
        return sum(a.criar for a in self.abas)

    @property
    def total_atualizar(self):
        return sum(a.atualizar for a in self.abas)

    @property
    def total_rejeitadas(self):
        return sum(len(a.rejeitadas) for a in self.abas)

    def resumo(self, nome_aba):
        for a in self.abas:
            if a.aba == nome_aba:
                return a
        return None


def _erro(mensagem):
    return Relatorio(abas=[ResumoAba(aba=n) for n in ABAS_NA_ORDEM], erro_estrutural=mensagem)


def _normalizar(valor):
    return str(valor).strip().lower() if valor is not None else ''


def _ler_abas(arquivo):
    """Devolve (abas, erro). abas = {nome: [(linha_excel, {coluna: valor})]}."""
    try:
        wb = load_workbook(arquivo, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001 -- openpyxl levanta tipos variados
        return None, 'Arquivo não pôde ser lido como .xlsx.'

    presentes = [n for n in ABAS_NA_ORDEM if n in wb.sheetnames]
    if not presentes:
        return None, 'Nenhuma aba reconhecida. Esperadas: ' + ', '.join(ABAS_NA_ORDEM) + '.'

    abas = {n: [] for n in ABAS_NA_ORDEM}
    for nome in presentes:
        ws = wb[nome]
        linhas = ws.iter_rows(values_only=True)
        try:
            cabecalho = [_normalizar(c) for c in next(linhas)]
        except StopIteration:
            continue  # aba presente mas totalmente vazia: nada a importar
        faltando = [c for c in COLUNAS_POR_ABA[nome] if c not in cabecalho]
        if faltando:
            return None, (
                f"Aba '{nome}': coluna(s) ausente(s) no cabeçalho: {', '.join(faltando)}."
            )
        for numero, valores in enumerate(linhas, start=2):
            if all(v is None for v in valores):
                continue
            abas[nome].append((numero, dict(zip(cabecalho, valores))))
    return abas, None


def _numero(valores, coluna):
    """Devolve (numero, erro). Célula em branco e valor não numérico são erro."""
    bruto = valores.get(coluna)
    if bruto is None or (isinstance(bruto, str) and not bruto.strip()):
        return None, f'{coluna} em branco'
    try:
        return float(bruto), None
    except (TypeError, ValueError):
        return None, f'{coluna} não é um número: {bruto!r}'


def _texto(valores, coluna):
    bruto = valores.get(coluna)
    if bruto is None:
        return None
    return str(bruto).strip() or None


def _analisar_unidades(aba, linhas, existentes, resumo):
    """Fábricas e Armazéns: mesma forma, só muda a lista de obrigatórios.

    Devolve o conjunto de nomes que a pasta vai criar ou atualizar, para as
    abas seguintes resolverem contra ele.
    """
    nomes = set()
    max_nome = MODELO_POR_ABA[aba]._meta.get_field('nome').max_length
    for numero, valores in linhas:
        nome = _texto(valores, 'nome')
        if not nome:
            resumo.rejeitadas.append(LinhaRejeitada(aba, numero, 'nome em branco', valores))
            continue
        if len(nome) > max_nome:
            resumo.rejeitadas.append(LinhaRejeitada(
                aba, numero, f'nome tem mais de {max_nome} caracteres', valores,
            ))
            continue
        if nome in nomes:
            resumo.rejeitadas.append(LinhaRejeitada(aba, numero, 'nome duplicado na planilha', valores))
            continue
        erros = []
        for coluna in OBRIGATORIOS_POR_ABA[aba]:
            _, erro = _numero(valores, coluna)
            if erro:
                erros.append(erro)
        if erros:
            resumo.rejeitadas.append(LinhaRejeitada(aba, numero, '; '.join(erros), valores))
            continue
        if nome in existentes:
            resumo.atualizar += 1
        else:
            resumo.criar += 1
        nomes.add(nome)
    return nomes


def _data(valores, coluna):
    """Devolve (date, erro). Aceita date/datetime do Excel e texto ISO."""
    bruto = valores.get(coluna)
    if bruto is None or (isinstance(bruto, str) and not bruto.strip()):
        return None, f'{coluna} em branco'
    if isinstance(bruto, datetime.datetime):
        return bruto.date(), None
    if isinstance(bruto, datetime.date):
        return bruto, None
    try:
        return datetime.date.fromisoformat(str(bruto).strip()), None
    except ValueError:
        return None, f'{coluna} não é uma data válida (use AAAA-MM-DD): {bruto!r}'


def _resolver(nome, fabricas, armazens):
    """Devolve (tipo, erro). tipo é 'Fábrica' ou 'Armazém'.

    Nome presente nos dois conjuntos é ambíguo -- rejeita em vez de escolher.
    O legado escolhe a fábrica em silêncio (data_loader.py:439-453).
    """
    eh_fabrica = nome in fabricas
    eh_armazem = nome in armazens
    if eh_fabrica and eh_armazem:
        return None, (
            f"'{nome}' é ambíguo: existe uma fábrica e um armazém com esse nome neste cenário"
        )
    if eh_fabrica:
        return 'Fábrica', None
    if eh_armazem:
        return 'Armazém', None
    return None, f"'{nome}' não corresponde a nenhuma fábrica nem armazém deste cenário"


def _chaves_de_rota(cenario):
    if cenario is None:
        return set()
    return {
        (r.armazem.nome, r.fabrica.nome)
        for r in Rota.all_cooperativas.filter(cenario=cenario).select_related(
            'armazem', 'fabrica'
        )
    }


def _chaves_de_previsao(cenario):
    if cenario is None:
        return set()
    chaves = {
        ('Fábrica', p.fabrica.nome, p.mes_referencia)
        for p in PrevisaoFabrica.all_cooperativas.filter(
            fabrica__cenario=cenario
        ).select_related('fabrica')
    }
    chaves |= {
        ('Armazém', p.armazem.nome, p.mes_referencia)
        for p in PrevisaoArmazem.all_cooperativas.filter(
            armazem__cenario=cenario
        ).select_related('armazem')
    }
    return chaves


def _tipo_canonico(entidade_tipo):
    """'Armazém' identifica um armazém; qualquer outra grafia persistida
    (ex.: 'fabrica' minúsculo -- ver tests/test_models_a11.py e a nota em
    `_gravar_safras`) conta como fábrica. Extraído para que `_chaves_de_safra`
    (o que `analisar` promete) e `_gravar_safras` (o que `aplicar` grava)
    nunca divirjam sobre o que é canônico -- finding 4 do review da Task 3.
    """
    return 'Armazém' if entidade_tipo == 'Armazém' else 'Fábrica'


def _chaves_de_safra(cenario):
    if cenario is None:
        return set()
    fabricas = dict(
        Fabrica.all_cooperativas.filter(cenario=cenario).values_list('id', 'nome')
    )
    armazens = dict(
        Armazem.all_cooperativas.filter(cenario=cenario).values_list('id', 'nome')
    )
    chaves = set()
    for s in SafraUnidade.all_cooperativas.filter(cenario=cenario):
        tipo = _tipo_canonico(s.entidade_tipo)
        mapa = armazens if tipo == 'Armazém' else fabricas
        nome = mapa.get(s.entidade_id)
        if nome:
            chaves.add((tipo, nome, s.data_inicio))
    return chaves


def _analisar_rotas(linhas, fabricas, armazens, chaves_no_banco, resumo):
    vistos = set()
    for numero, valores in linhas:
        origem = _texto(valores, 'origem')
        destino = _texto(valores, 'destino')
        if not origem or not destino:
            resumo.rejeitadas.append(
                LinhaRejeitada(ABA_ROTAS, numero, 'origem ou destino em branco', valores)
            )
            continue
        if origem not in armazens:
            resumo.rejeitadas.append(LinhaRejeitada(
                ABA_ROTAS, numero,
                f"origem '{origem}' não corresponde a nenhum armazém deste cenário", valores,
            ))
            continue
        if destino not in fabricas:
            resumo.rejeitadas.append(LinhaRejeitada(
                ABA_ROTAS, numero,
                f"destino '{destino}' não corresponde a nenhuma fábrica deste cenário", valores,
            ))
            continue
        if (origem, destino) in vistos:
            resumo.rejeitadas.append(
                LinhaRejeitada(ABA_ROTAS, numero, 'rota duplicada na planilha', valores)
            )
            continue
        erros = []
        for coluna in ('distancia_km', 'custo_frete_ton'):
            _, erro = _numero(valores, coluna)
            if erro:
                erros.append(erro)
        # custo_frete_entressafra em branco assume custo_frete_ton
        # (data_loader.py:386-387); só um valor presente e não numérico é erro.
        if valores.get('custo_frete_entressafra') is not None:
            _, erro = _numero(valores, 'custo_frete_entressafra')
            if erro:
                erros.append(erro)
        if erros:
            resumo.rejeitadas.append(
                LinhaRejeitada(ABA_ROTAS, numero, '; '.join(erros), valores)
            )
            continue
        if (origem, destino) in chaves_no_banco:
            resumo.atualizar += 1
        else:
            resumo.criar += 1
        vistos.add((origem, destino))


def _analisar_previsoes(linhas, fabricas, armazens, chaves_no_banco, resumo):
    vistos = set()
    for numero, valores in linhas:
        nome = _texto(valores, 'entidade')
        if not nome:
            resumo.rejeitadas.append(
                LinhaRejeitada(ABA_PREVISOES, numero, 'entidade em branco', valores)
            )
            continue
        tipo, erro = _resolver(nome, fabricas, armazens)
        if erro:
            resumo.rejeitadas.append(LinhaRejeitada(ABA_PREVISOES, numero, erro, valores))
            continue
        mes, erro = _data(valores, 'mes_referencia')
        if erro:
            resumo.rejeitadas.append(LinhaRejeitada(ABA_PREVISOES, numero, erro, valores))
            continue
        mes = mes.replace(day=1)
        if (nome, mes) in vistos:
            resumo.rejeitadas.append(LinhaRejeitada(
                ABA_PREVISOES, numero,
                'previsão duplicada na planilha (mesma entidade e mês)', valores,
            ))
            continue
        # recebimento_produtor e vendas em branco valem 0 (bug A9 da Fase 1).
        erro_numerico = None
        for coluna in ('recebimento_produtor', 'vendas'):
            if valores.get(coluna) is None:
                continue
            _, erro = _numero(valores, coluna)
            if erro:
                erro_numerico = erro
                break
        if erro_numerico:
            resumo.rejeitadas.append(
                LinhaRejeitada(ABA_PREVISOES, numero, erro_numerico, valores)
            )
            continue
        if (tipo, nome, mes) in chaves_no_banco:
            resumo.atualizar += 1
        else:
            resumo.criar += 1
        vistos.add((nome, mes))


def _analisar_safras(linhas, fabricas, armazens, chaves_no_banco, resumo):
    vistos = set()
    for numero, valores in linhas:
        nome = _texto(valores, 'unidade')
        if not nome:
            resumo.rejeitadas.append(
                LinhaRejeitada(ABA_SAFRAS, numero, 'unidade em branco', valores)
            )
            continue
        tipo, erro = _resolver(nome, fabricas, armazens)
        if erro:
            resumo.rejeitadas.append(LinhaRejeitada(ABA_SAFRAS, numero, erro, valores))
            continue
        inicio, erro_i = _data(valores, 'data_inicio')
        fim, erro_f = _data(valores, 'data_fim')
        if erro_i or erro_f:
            resumo.rejeitadas.append(LinhaRejeitada(
                ABA_SAFRAS, numero, '; '.join(e for e in (erro_i, erro_f) if e), valores,
            ))
            continue
        if fim < inicio:
            resumo.rejeitadas.append(LinhaRejeitada(
                ABA_SAFRAS, numero, 'data_fim é anterior a data_inicio', valores,
            ))
            continue
        if (nome, inicio) in vistos:
            resumo.rejeitadas.append(LinhaRejeitada(
                ABA_SAFRAS, numero,
                'safra duplicada na planilha (mesma unidade e data de início)', valores,
            ))
            continue
        if (tipo, nome, inicio) in chaves_no_banco:
            resumo.atualizar += 1
        else:
            resumo.criar += 1
        vistos.add((nome, inicio))


def analisar(arquivo, cenario):
    """Lê a pasta e classifica cada linha. NUNCA escreve.

    `cenario` pode ser None (bootstrap: o cenário ainda não existe). Nesse caso
    o lado-banco da resolução é vazio -- tudo é criação, e os nomes resolvem
    apenas contra a própria pasta.
    """
    abas, erro = _ler_abas(arquivo)
    if erro:
        return _erro(erro)

    if cenario is not None:
        fabricas_no_banco = set(
            Fabrica.all_cooperativas.filter(cenario=cenario).values_list('nome', flat=True)
        )
        armazens_no_banco = set(
            Armazem.all_cooperativas.filter(cenario=cenario).values_list('nome', flat=True)
        )
    else:
        fabricas_no_banco = set()
        armazens_no_banco = set()

    relatorio = Relatorio(abas=[ResumoAba(aba=n) for n in ABAS_NA_ORDEM])

    nomes_fabricas = fabricas_no_banco | _analisar_unidades(
        ABA_FABRICAS, abas[ABA_FABRICAS], fabricas_no_banco, relatorio.resumo(ABA_FABRICAS),
    )
    nomes_armazens = armazens_no_banco | _analisar_unidades(
        ABA_ARMAZENS, abas[ABA_ARMAZENS], armazens_no_banco, relatorio.resumo(ABA_ARMAZENS),
    )

    _analisar_rotas(
        abas[ABA_ROTAS], nomes_fabricas, nomes_armazens,
        _chaves_de_rota(cenario), relatorio.resumo(ABA_ROTAS),
    )
    _analisar_previsoes(
        abas[ABA_PREVISOES], nomes_fabricas, nomes_armazens,
        _chaves_de_previsao(cenario), relatorio.resumo(ABA_PREVISOES),
    )
    _analisar_safras(
        abas[ABA_SAFRAS], nomes_fabricas, nomes_armazens,
        _chaves_de_safra(cenario), relatorio.resumo(ABA_SAFRAS),
    )

    return relatorio


def aplicar(arquivo, cenario=None, cooperativa=None, nome_novo=None):
    """Reanalisa a pasta e grava as linhas válidas. Devolve (relatorio, cenario).

    Com `cenario=None`, cria o cenário a partir de `cooperativa` e `nome_novo`,
    marcando-o oficial se for o primeiro daquela cooperativa. A criação acontece
    DENTRO da mesma transação que grava as linhas, para que uma pasta com erro
    não deixe um cenário vazio para trás.

    Devolve (relatorio, None) sem escrever nada quando há erro estrutural.

    Cada `_gravar_*` espelha as rejeições do `_analisar_*` correspondente
    (nome/chave em branco, referência não resolvida, duplicata na própria
    planilha, campo numérico/data inválido) para que o relatório devolvido
    por `analisar` nunca prometa uma escrita que `aplicar` não faz.
    """
    relatorio = analisar(arquivo, cenario)
    if relatorio.tem_erro_estrutural:
        return relatorio, None

    arquivo.seek(0)
    abas, erro = _ler_abas(arquivo)
    if erro:  # defensivo: a mesma pasta acabou de ser lida com sucesso
        return _erro(erro), None

    with transaction.atomic():
        if cenario is None:
            primeiro = not Cenario.all_cooperativas.filter(cooperativa=cooperativa).exists()
            cenario = Cenario(
                cooperativa=cooperativa, nome=nome_novo, is_oficial=primeiro,
            )
            cenario.full_clean()
            cenario.save()
        coop = cenario.cooperativa

        fabricas = _gravar_unidades(Fabrica, ABA_FABRICAS, abas[ABA_FABRICAS], cenario, coop)
        armazens = _gravar_unidades(Armazem, ABA_ARMAZENS, abas[ABA_ARMAZENS], cenario, coop)
        _gravar_rotas(abas[ABA_ROTAS], cenario, coop, fabricas, armazens)
        _gravar_previsoes(abas[ABA_PREVISOES], coop, fabricas, armazens)
        _gravar_safras(abas[ABA_SAFRAS], cenario, coop, fabricas, armazens)

    return relatorio, cenario


def _gravar_unidades(modelo, aba, linhas, cenario, coop):
    """Fábricas e Armazéns. Devolve {nome: instância} do cenário ao final.

    Espelha `_analisar_unidades`: nome em branco, nome maior que o
    `max_length` do model (finding 2 do review da Task 3), e uma segunda
    ocorrência do mesmo nome na planilha (correção a) são pulados -- gravar
    qualquer um deles sobrescreveria/violaria o que `analisar` já classificou.
    """
    existentes = {u.nome: u for u in modelo.all_cooperativas.filter(cenario=cenario)}
    max_nome = modelo._meta.get_field('nome').max_length
    vistos = set()
    for _, valores in linhas:
        nome = _texto(valores, 'nome')
        if not nome:
            continue
        if len(nome) > max_nome:
            continue
        if nome in vistos:
            continue
        numeros = {}
        invalida = False
        for coluna in OBRIGATORIOS_POR_ABA[aba]:
            valor, erro = _numero(valores, coluna)
            if erro:
                invalida = True
                break
            numeros[coluna] = valor
        if invalida:
            continue
        vistos.add(nome)
        unidade = existentes.get(nome) or modelo(cooperativa=coop, cenario=cenario, nome=nome)
        for coluna, valor in numeros.items():
            setattr(unidade, coluna, valor)
        if aba == ABA_FABRICAS:
            unidade.limite_caminhoes = int(unidade.limite_caminhoes)
        unidade.full_clean()
        unidade.save()
        existentes[nome] = unidade
    return existentes


def _gravar_rotas(linhas, cenario, coop, fabricas, armazens):
    """Espelha `_analisar_rotas`: origem/destino não resolvidos e uma segunda
    ocorrência da mesma (origem, destino) na planilha são puladas (correção b).
    """
    vistos = set()
    for _, valores in linhas:
        origem = _texto(valores, 'origem')
        destino = _texto(valores, 'destino')
        if origem not in armazens or destino not in fabricas:
            continue
        if (origem, destino) in vistos:
            continue
        distancia, erro_d = _numero(valores, 'distancia_km')
        custo, erro_c = _numero(valores, 'custo_frete_ton')
        if erro_d or erro_c:
            continue
        if valores.get('custo_frete_entressafra') is None:
            entressafra = custo
        else:
            entressafra, erro_e = _numero(valores, 'custo_frete_entressafra')
            if erro_e:
                continue
        vistos.add((origem, destino))
        armazem = armazens[origem]
        fabrica = fabricas[destino]
        rota = Rota.all_cooperativas.filter(
            cenario=cenario, armazem=armazem, fabrica=fabrica,
        ).first() or Rota(
            cooperativa=coop, cenario=cenario, armazem=armazem, fabrica=fabrica,
        )
        rota.distancia_km = distancia
        rota.custo_frete_ton = custo
        rota.custo_frete_entressafra = entressafra
        rota.full_clean()
        rota.save()


def _gravar_previsoes(linhas, coop, fabricas, armazens):
    """Espelha `_analisar_previsoes`: entidade não resolvida, mês não
    parseável e uma segunda ocorrência de (nome, mês normalizado ao dia 1)
    na planilha são puladas (correção b).
    """
    vistos = set()
    for _, valores in linhas:
        nome = _texto(valores, 'entidade')
        if not nome:
            continue
        tipo, erro = _resolver(nome, set(fabricas), set(armazens))
        if erro:
            continue
        mes, erro = _data(valores, 'mes_referencia')
        if erro:
            continue
        mes = mes.replace(day=1)
        if (nome, mes) in vistos:
            continue
        numeros = {}
        invalida = False
        for coluna in ('recebimento_produtor', 'vendas'):
            if valores.get(coluna) is None:
                numeros[coluna] = 0
                continue
            valor, erro = _numero(valores, coluna)
            if erro:
                invalida = True
                break
            numeros[coluna] = valor
        if invalida:
            continue
        vistos.add((nome, mes))
        if tipo == 'Fábrica':
            modelo, campo, unidade = PrevisaoFabrica, 'fabrica', fabricas[nome]
        else:
            modelo, campo, unidade = PrevisaoArmazem, 'armazem', armazens[nome]
        previsao = modelo.all_cooperativas.filter(
            mes_referencia=mes, **{campo: unidade},
        ).first() or modelo(cooperativa=coop, mes_referencia=mes, **{campo: unidade})
        previsao.recebimento_produtor = numeros['recebimento_produtor']
        previsao.vendas = numeros['vendas']
        previsao.full_clean()
        previsao.save()


def _gravar_safras(linhas, cenario, coop, fabricas, armazens):
    """Espelha `_analisar_safras`: unidade não resolvida, datas inválidas e
    uma segunda ocorrência de (nome, data_inicio) na planilha são puladas
    (correção b).

    Correção (c): `entidade_tipo` no banco nem sempre é grafado de forma
    canônica (ver `_chaves_de_safra` -- variantes como 'fabrica' minúsculo
    se propagam por cenários clonados e importados do legado). Por isso a
    consulta NÃO filtra por `entidade_tipo`: filtra só por
    (cenario, entidade_id, data_inicio) e escolhe, entre os resultados, o
    que canonicaliza para o tipo desta linha -- `entidade_id` sozinho não
    identifica a unidade porque fábricas e armazéns vêm de sequências de id
    independentes e podem colidir. Ao gravar, `entidade_tipo` é sempre
    ajustado para o valor canônico, convergindo o dado ao longo do tempo.
    """
    vistos = set()
    for _, valores in linhas:
        nome = _texto(valores, 'unidade')
        if not nome:
            continue
        tipo, erro = _resolver(nome, set(fabricas), set(armazens))
        if erro:
            continue
        inicio, erro_i = _data(valores, 'data_inicio')
        fim, erro_f = _data(valores, 'data_fim')
        if erro_i or erro_f or fim < inicio:
            continue
        if (nome, inicio) in vistos:
            continue
        vistos.add((nome, inicio))
        unidade = armazens[nome] if tipo == 'Armazém' else fabricas[nome]
        candidatas = SafraUnidade.all_cooperativas.filter(
            cenario=cenario, entidade_id=unidade.id, data_inicio=inicio,
        )
        safra = next(
            (s for s in candidatas if _tipo_canonico(s.entidade_tipo) == tipo),
            None,
        ) or SafraUnidade(
            cooperativa=coop, cenario=cenario, entidade_id=unidade.id, data_inicio=inicio,
        )
        safra.entidade_tipo = tipo
        safra.data_fim = fim
        safra.full_clean()
        safra.save()
